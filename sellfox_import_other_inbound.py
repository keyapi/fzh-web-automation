#!/usr/bin/env python3
"""
sellfox_import_other_inbound.py
赛狐 其他入库 导入 + 自验证 — 参考 sellfox_import_update.py 的闭环验证模式

用法:
  uv run python sellfox_import_other_inbound.py                         # 使用默认测试数据
  uv run python sellfox_import_other_inbound.py --file test.xlsx         # 导入指定文件
  uv run python sellfox_import_other_inbound.py --sku test001-white --wh POLAND --qty 1000 --price 1.0
  uv run python sellfox_import_other_inbound.py --headless               # 无头模式

流程: 生成Excel → 导航→添加单据→导入入库单→上传→导入→等结果→自验证
"""

import sys
import time
import shutil
from pathlib import Path
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

import pandas as pd
import openpyxl
from playwright.sync_api import sync_playwright

SCRIPT_DIR = Path(__file__).resolve().parent
PROFILE_DIR = SCRIPT_DIR / "sellfox-profile"
DOWNLOADS_DIR = SCRIPT_DIR / "downloads"
TEMPLATE_FILE = SCRIPT_DIR / ".playwright-mcp" / "入库单导入模板.xlsx"

PAGE_URL = "https://www.sellfox.com/amzup-web-main/web/warehouse/otherIn/index.html"
LOGIN_URL = "https://www.sellfox.com/amzup-web-main/login.html"
LOGIN_TIMEOUT = 300

INBOUND_COLUMNS = [
    "临时单号", "*入库仓库", "*入库类型", "运费（CNY）", "其他费用（CNY）",
    "费用分摊方式", "单位费用", "单据备注", "入库时间", "*SKU",
    "店铺", "FNSKU", "专属类型", "*采购单价(CNY)", "可用数",
    "次品数", "可用货架位", "次品货架位", "商品备注",
]

# ── 工具函数 ──────────────────────────────────────────────

def wait_for_login(page):
    print(f"\n请在浏览器中登录赛狐（最长等待 {LOGIN_TIMEOUT}s）...")
    for i in range(0, LOGIN_TIMEOUT, 2):
        time.sleep(2)
        url = page.url
        if "login" not in url and url != "about:blank" and "sellfox" in url:
            print(f"✓ 检测到登录成功！")
            page.wait_for_timeout(1000)
            return True
        try:
            if page.locator('text=克勇').first.is_visible():
                print("✓ 检测到登录成功！(用户菜单可见)")
                return True
        except:
            pass
    return False


def is_logged_in(page):
    url = page.url
    if "login" in url or url == "about:blank":
        return False
    try:
        return page.locator("button:has-text(\"添加单据\")").first.is_visible(timeout=3000)
    except:
        return "login" not in url


# ── Excel 生成 ────────────────────────────────────────────

def make_inbound_excel(sku: str, warehouse: str, qty: int, price: float,
                       out_path: Path, note: str = ""):
    """生成其他入库导入 Excel（保留模板 Data Validation sheet）。"""
    if TEMPLATE_FILE.exists():
        shutil.copy(TEMPLATE_FILE, out_path)
        wb = openpyxl.load_workbook(out_path)
        ws = wb["sheet1"]
    else:
        # Fallback: 从零创建（无 Data Validation，可能被赛狐拒绝）
        print("  [!] 模板不存在，从零创建（可能被赛狐拒绝）")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "sheet1"
        for i, col_name in enumerate(INBOUND_COLUMNS, 1):
            ws.cell(row=1, column=i, value=col_name)

    row = 2  # 数据从第2行开始（1行表头）
    ws.cell(row=row, column=2, value=warehouse)   # *入库仓库
    ws.cell(row=row, column=3, value="其他入库")   # *入库类型
    ws.cell(row=row, column=10, value=sku)         # *SKU
    ws.cell(row=row, column=14, value=price)       # *采购单价(CNY)
    ws.cell(row=row, column=15, value=qty)          # 可用数
    if note:
        ws.cell(row=row, column=8, value=note)     # 单据备注

    wb.save(out_path)
    return out_path


# ── 导入 + 自验证 ────────────────────────────────────────

def import_and_verify(page, filepath: Path, sku: str, warehouse: str, expected_qty: int) -> dict:
    """导入一个入库文件并验证结果。"""
    print(f"\n{'='*50}")
    print(f"导入其他入库: {filepath.name}")
    print(f"  SKU={sku}, 仓库={warehouse}, 数量={expected_qty}")
    print(f"{'='*50}")

    # ── 导航 ──
    print("  导航到其他入库页...")
    page.goto(PAGE_URL, timeout=30000)
    page.wait_for_timeout(3000)
    page.keyboard.press("Escape")
    page.wait_for_timeout(300)

    # ── 添加单据 → 导入入库单 ──
    print("  添加单据 → 导入入库单...")
    page.evaluate("""
        (() => {
            for (const el of document.querySelectorAll('button')) {
                if ((el.textContent || '').includes('添加单据') && el.offsetParent !== null) {
                    el.click();
                    break;
                }
            }
        })()
    """)
    page.wait_for_timeout(600)
    page.evaluate("""
        (() => {
            for (const item of document.querySelectorAll('.el-dropdown-menu__item')) {
                if ((item.textContent || '').includes('导入入库单')) {
                    item.click();
                    return;
                }
            }
        })()
    """)
    page.wait_for_timeout(1500)

    # ── 上传文件 ──
    print("  上传文件...")
    add_file = page.locator("button").filter(has_text="添加文件").first
    with page.expect_file_chooser(timeout=10000) as fc_info:
        add_file.click(timeout=5000)
    file_chooser = fc_info.value
    page.wait_for_timeout(500)
    file_chooser.set_files(str(filepath))
    page.wait_for_timeout(1000)

    # ── 导入 ──
    print("  导入中...")
    page.evaluate("""
        (() => {
            const wrappers = document.querySelectorAll('.el-dialog__wrapper');
            for (const d of wrappers) {
                if (window.getComputedStyle(d).display !== 'none') {
                    for (const btn of d.querySelectorAll('button')) {
                        if (btn.textContent.trim() === '导入' && !btn.disabled) {
                            btn.click();
                            return;
                        }
                    }
                }
            }
        })()
    """)

    # ── 等待结果 ──
    result_text = None
    for _ in range(40):
        time.sleep(2)
        result = page.evaluate("""
            (() => {
                const wrappers = document.querySelectorAll('.el-dialog__wrapper');
                for (const d of wrappers) {
                    if (window.getComputedStyle(d).display !== 'none') {
                        const t = d.textContent || '';
                        if (t.includes('导入完成')) return t;
                    }
                }
                return null;
            })()
        """)
        if result:
            result_text = result
            break

    if not result_text:
        print("  [!] 超时未收到导入结果")
        return {"success": False, "error": "timeout"}

    import re
    success_match = re.search(r'成功(\d+)条', result_text)
    fail_match = re.search(r'失败(\d+)条', result_text)
    success_count = int(success_match.group(1)) if success_match else 0
    fail_count = int(fail_match.group(1)) if fail_match else 0
    print(f"  导入结果: 成功{success_count}条, 失败{fail_count}条")

    # ── 关闭弹窗 ──
    page.evaluate("""
        document.querySelectorAll('.el-dialog__wrapper button').forEach(btn => {
            if (btn.textContent.trim() === '关闭') btn.click();
        });
    """)
    page.wait_for_timeout(500)

    if fail_count > 0:
        return {"success": False, "success_count": success_count, "fail_count": fail_count}

    # ── 自验证：刷新页面，搜索 SKU，确认记录存在 ──
    print("  自验证: 搜索 SKU 确认入库记录...")
    page.goto(PAGE_URL, timeout=30000)
    page.wait_for_timeout(5000)
    page.keyboard.press("Escape")
    page.wait_for_timeout(300)

    # 先切换搜索类型为 SKU（参考 sellfox_auto_export.py 的 switch_search_type）
    page.evaluate("""
        (() => {
            const inputs = document.querySelectorAll('input.el-input__inner');
            for (const inp of inputs) {
                const v = inp.value;
                if (!v) continue;
                if (['SKU','识别码','品名','型号','FNSKU','SPU','款名','MSKU','入库单号','备货单号'].includes(v)) {
                    if (v !== 'SKU') {
                        const sel = inp.closest('.el-select');
                        if (sel) sel.click();
                    }
                    return;
                }
            }
        })()
    """)
    page.wait_for_timeout(500)
    # 点击 SKU 选项
    page.evaluate("""
        (() => {
            const items = [...document.querySelectorAll('.el-select-dropdown__item')]
                .filter(i => i.getBoundingClientRect().width > 0);
            const m = items.find(i => i.textContent.trim() === 'SKU');
            if (m) m.click();
        })()
    """)
    page.wait_for_timeout(300)

    # 搜索 SKU
    search_input = page.locator("input.el-input__inner[placeholder='搜索内容']")
    if search_input.count() == 0:
        search_input = page.locator("input[placeholder='搜索内容']").first
    search_input.fill(sku)
    search_input.press("Enter")
    page.wait_for_timeout(3000)

    # 检查结果
    verify = page.evaluate(f"""
        (() => {{
            const body = document.body.innerText;
            const hasSku = body.includes('{sku}');
            const hasWarehouse = body.includes('{warehouse}');
            const hasQty = body.includes('{expected_qty}');

            // 检查分页
            const pagination = document.querySelector('.el-pagination');
            const pagText = pagination ? pagination.textContent.trim() : '';

            // 检查表格行
            const rows = document.querySelectorAll('.vxe-table--body tbody tr');
            let skuFound = false;
            rows.forEach(r => {{
                if (r.textContent.includes('{sku}')) skuFound = true;
            }});

            return {{ hasSku, hasWarehouse, hasQty, skuFound, rowCount: rows.length, pagText }};
        }})()
    """)

    verified = verify.get("skuFound", False)
    if verified:
        print(f"  ✓ 验证通过: SKU={sku} 出现在列表中 (共{verify.get('rowCount', 0)}行)")
    else:
        print(f"  ✗ 验证失败: SKU={sku} 未找到! 分页={verify.get('pagText', '')}")

    # ── 确认入库 ──
    confirmed = False
    if verified:
        print("  确认入库...")
        # 可能有多个待确认记录（之前测试遗留），逐个点击确认
        confirm_clicked = 0
        for attempt in range(10):
            clicked = page.evaluate("""
                (() => {
                    const allBtns = document.querySelectorAll('button');
                    for (const btn of allBtns) {
                        if (btn.textContent.trim() === '确认入库' && btn.offsetParent !== null) {
                            btn.click();
                            return 'clicked';
                        }
                    }
                    return 'done';
                })()
            """)
            if clicked == 'done':
                break
            confirm_clicked += 1
            page.wait_for_timeout(1000)
        print(f"  已点击 {confirm_clicked} 个确认入库按钮")
        if confirm_clicked > 0:
            print("  等待确认生效 + 刷新验证...")
            page.wait_for_timeout(3000)
            page.goto(PAGE_URL, timeout=30000)
            page.wait_for_timeout(5000)
            page.keyboard.press("Escape")
            page.wait_for_timeout(300)
            # 切 SKU + 搜索
            page.evaluate("""
                (() => { const inputs = document.querySelectorAll('input.el-input__inner');
                  for (const inp of inputs) { const v = inp.value;
                    if (v && ['SKU','识别码','品名','型号','FNSKU','SPU','款名','MSKU','入库单号'].includes(v)) {
                      if (v !== 'SKU') { const sel = inp.closest('.el-select'); if (sel) sel.click(); } return; } } })()
            """)
            page.wait_for_timeout(500)
            page.evaluate("""
                (() => { const items = [...document.querySelectorAll('.el-select-dropdown__item')]
                  .filter(i => i.getBoundingClientRect().width > 0);
                  const m = items.find(i => i.textContent.trim() === 'SKU'); if (m) m.click(); })()
            """)
            page.wait_for_timeout(300)
            si = page.locator("input.el-input__inner[placeholder='搜索内容']")
            if si.count() == 0: si = page.locator("input[placeholder='搜索内容']").first
            si.fill(sku); si.press("Enter")
            page.wait_for_timeout(3000)
            remaining = page.evaluate("""
                (() => { let c = 0; document.querySelectorAll('button').forEach(b => {
                  if (b.textContent.trim() === '确认入库' && b.offsetParent !== null) c++; }); return c; })()
            """)
            confirmed = (remaining == 0)
            print(f"  {'✓ 全部确认完成' if confirmed else '⚠ 仍有' + str(remaining) + '条待确认'}")
        else:
            print(f"  ⚠ 未找到确认入库按钮 (可能已确认或行已折叠)")

    return {
        "success": True,
        "success_count": success_count,
        "fail_count": fail_count,
        "verified": verified,
        "confirmed": confirmed,
    }


# ── 主入口 ────────────────────────────────────────────────

def main():
    headless = "--headless" in sys.argv

    # 解析参数
    sku = "test001-white"
    warehouse = "POLAND"
    qty = 1000
    price = 1.0
    note = ""
    file_arg = None

    args = sys.argv[1:]
    i = 0
    while i < len(args):
        if args[i] == "--sku" and i + 1 < len(args):
            sku = args[i + 1]; i += 2
        elif args[i] == "--wh" and i + 1 < len(args):
            warehouse = args[i + 1]; i += 2
        elif args[i] == "--qty" and i + 1 < len(args):
            qty = int(args[i + 1]); i += 2
        elif args[i] == "--price" and i + 1 < len(args):
            price = float(args[i + 1]); i += 2
        elif args[i] == "--note" and i + 1 < len(args):
            note = args[i + 1]; i += 2
        elif args[i] == "--file" and i + 1 < len(args):
            file_arg = Path(args[i + 1]); i += 2
        elif args[i] == "--headless":
            i += 1
        else:
            i += 1

    # 生成 Excel
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if file_arg:
        filepath = file_arg
    else:
        filepath = DOWNLOADS_DIR / f"其他入库_{sku}_{warehouse}_{stamp}.xlsx"
        DOWNLOADS_DIR.mkdir(exist_ok=True)
        make_inbound_excel(sku, warehouse, qty, price, filepath, note)
        print(f"生成导入文件: {filepath.name}")

    if not filepath.exists():
        print(f"[失败] 文件不存在: {filepath}")
        return

    # 导入
    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE_DIR),
            headless=headless,
        )
        page = context.pages[0] if context.pages else context.new_page()

        page.goto(PAGE_URL, timeout=30000)
        page.wait_for_timeout(3000)

        if is_logged_in(page):
            print("\n[OK] 已登录，跳过登录步骤")
        else:
            print("\n[!] 未登录 → 打开登录页")
            page.goto(LOGIN_URL, timeout=30000)
            if not wait_for_login(page):
                print("[失败] 登录超时")
                context.close()
                return

        result = import_and_verify(page, filepath, sku, warehouse, qty)

        print(f"\n{'='*50}")
        print("结果")
        print(f"{'='*50}")
        print(f"  导入: {'✓' if result.get('success') else '✗'}")
        if result.get("verified") is not None:
            print(f"  验证: {'✓' if result['verified'] else '✗'}")
        context.close()


if __name__ == "__main__":
    main()
